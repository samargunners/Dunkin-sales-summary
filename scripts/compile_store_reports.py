# scripts/compile_store_reports.py

from pathlib import Path
import pandas as pd
import re
import unicodedata
from openpyxl import load_workbook

# =================== USER TWEAKS ===================
DEBUG = True                      # verbose console + log output
INCLUDE_SALES_MIX_DETAIL = True   # <— now ON (you asked not to miss it)
# ===================================================

# ── Paths
BASE_DIR = Path(__file__).resolve().parents[1]
RAW_DIR  = BASE_DIR / "data" / "raw_emails"
OUT_DIR  = BASE_DIR / "data" / "compiled"
LOG_DIR  = BASE_DIR / "logs"
OUT_DIR.mkdir(parents=True, exist_ok=True)
LOG_DIR.mkdir(parents=True, exist_ok=True)
LOG_FILE = LOG_DIR / "compile_log.txt"

def log(msg: str) -> None:
    if DEBUG:
        print(msg)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(f"{msg}\n")

def norm(s: str) -> str:
    if s is None: return ""
    return unicodedata.normalize("NFKC", str(s)).replace("\u00A0"," ").replace("\u2011","-").strip()

# ── PC → Store mapping (from your reference)
PC_TO_STORE = {
    "301290":"Paxton","343939":"MountJoy","357993":"Enola",
    "358529":"Columbia","359042":"Lititz","363271":"Marietta","364322":"Etown",
}
FALLBACK_STORE_TOKENS = {
    "paxton":"Paxton","enola":"Enola","columbia":"Columbia","lititz":"Lititz",
    "marietta":"Marietta","mount joy":"MountJoy","mountjoy":"MountJoy","etown":"Etown",
}

# ── Default schemas (used if no *_copy.xlsx template is found)
DEFAULT_SCHEMAS = {
    "sales_summary": [
        "Store","PC_Number","Date","Gross_Sales","Net_Sales",
        "DD_Adjusted_No_Markup","PA_Sales_Tax","DD_Discount","Guest_Count",
        "Avg_Check","Gift_Card_Sales","Void_Amount","Refund","Void_Qty",
        "Paid_IN","Paid_OUT","Cash_IN"
    ],
    "sales_by_daypart": [
        "Store","PC_Number","Date","Daypart","Net_Sales","Percent_Sales","Check_Count","Avg_Check"
    ],
    "sales_by_subcategory": [
        "Store","PC_Number","Date","Subcategory","Qty_Sold","Net_Sales","Percent_Sales"
    ],
    "tender_type_metrics": [
        "Store","PC_Number","Date","Tender_Type","Detail_Amount"
    ],
    "labor_metrics": [
        "Store","PC_Number","Date","Labor_Position","Reg_Hours","OT_Hours","Total_Hours",
        "Reg_Pay","OT_Pay","Total_Pay","Percent_Labor"
    ],
    "sales_by_order_type": [
        "Store","PC_Number","Date","Order_Type","Net_Sales","Percent_Sales","Guests","Percent_Guest","Avg_Check"
    ],
    # reasonable default for sales mix detail (item-level)
    "sales_mix_detail": [
        "Store","PC_Number","Date","Category","Subcategory","Item","Qty_Sold","Net_Sales","Percent_Sales","Avg_Check"
    ],
}

# ── Utilities
def clean_num(x):
    if x is None or (isinstance(x,float) and pd.isna(x)): return pd.NA
    if isinstance(x,(int,float)): return float(x)
    s = norm(x)
    if s in {"","--","—","–"}: return pd.NA
    s = re.sub(r"[^\d.\-]","",s)
    if not re.search(r"\d",s): return pd.NA
    try: return float(s)
    except ValueError: return pd.NA

def find_col(df: pd.DataFrame, patterns):
    cols_norm = [re.sub(r"\s+"," ", norm(str(c))).lower() for c in df.columns]
    for pat in patterns:
        rx = re.compile(pat, re.I)
        for i,c in enumerate(cols_norm):
            if rx.search(c):
                return df.columns[i]
    return None

def filename_date_fallback(filename: str):
    m = re.search(r"(\d{4}-\d{2}-\d{2})", filename)
    if m:
        try: return pd.to_datetime(m.group(1)).date()
        except Exception: return pd.NaT
    return pd.NaT

def wb_sheetnames(path: Path):
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception as e:
        log(f"ERROR opening workbook {path.name}: {e}")
        return []

HEADER_KEYWORDS = {
    "pc","pc number","store #","store number","restaurant number",
    "store","store name","restaurant","location","location name","revenue center","site","unit","unit name",
    "date","business date","report date","period","day",
    "net","sales","gross","avg","check","checks","qty","quantity",
    "subcategory","category","order type","daypart","tender","position","labor","item","product",
    "amount","detail amount","percent","%"
}

def normalize_and_dedupe_columns(df: pd.DataFrame) -> pd.DataFrame:
    cols = [re.sub(r"\s+"," ", norm(str(c))).strip() for c in df.columns]
    out = df.copy()
    out.columns = cols
    to_drop = [i for i,c in enumerate(out.columns) if c == "" or c.lower() == "nan"]
    if to_drop:
        out = out.drop(out.columns[to_drop], axis=1)
    out = out.loc[:, ~pd.Index(out.columns).duplicated(keep="first")]
    return out

def read_with_header_detection(path: Path, sheet, extra_terms=None):
    df0 = pd.read_excel(path, sheet_name=sheet, header=None, engine="openpyxl")
    df0 = df0.dropna(how="all")
    if df0.empty: return None, None, None, None
    extra_terms = set(t.lower() for t in (extra_terms or []))
    def score(values):
        sc = 0
        for v in values:
            s = norm(v).lower()
            for k in (HEADER_KEYWORDS | extra_terms):
                if k in s and len(s) <= 64:
                    sc += 1; break
        return sc
    scan = min(10, len(df0))
    best_idx, best_score = 0, -1
    for i in range(scan):
        sc = score(list(df0.iloc[i].values))
        if sc > best_score:
            best_idx, best_score = i, sc
    df = df0.iloc[best_idx+1:].copy()
    df.columns = df0.iloc[best_idx].astype(str).tolist()
    df = df.dropna(how="all")
    df = normalize_and_dedupe_columns(df)
    return df, best_idx, best_score, df0

def category_from_filename(name_low: str):
    name_low = re.sub(r"\s+"," ", name_low)
    if "menu mix metrics" in name_low or "sales mix metrics" in name_low: return "sales_summary"
    if "sales by daypart"  in name_low: return "sales_by_daypart"
    if "sales by subcategory" in name_low: return "sales_by_subcategory"
    if "tender type" in name_low: return "tender_type_metrics"
    if "labor hours" in name_low or "labor metrics" in name_low: return "labor_metrics"
    if "order type" in name_low: return "sales_by_order_type"
    if "sales mix detail" in name_low: return "sales_mix_detail" if INCLUDE_SALES_MIX_DETAIL else None
    return None

PREFER_KEYS = {
    "sales_summary": ["menu","mix","metrics","sales mix"],
    "sales_by_daypart": ["daypart","part"],
    "sales_by_subcategory": ["subcategory","subcat"],
    "tender_type_metrics": ["tender","tender type"],
    "labor_metrics": ["labor","hours","metrics"],
    "sales_by_order_type": ["order type","order"],
    "sales_mix_detail": ["sales mix detail","mix detail","item"],
}
def prefer_sheets_for_category(sheetnames, category):
    keys = PREFER_KEYS.get(category, [])
    scored = []
    for s in sheetnames:
        sn = norm(s).lower()
        scored.append((1 if any(k in sn for k in keys) else 0, s))
    scored.sort(key=lambda x: -x[0])
    return [s for _, s in scored] or list(sheetnames)

# ── PC/Store/Date derivation
PC_PAT = re.compile(r"^\s*(\d{6})\b")
def derive_pc_store_from_text(txt: str):
    s = norm(txt)
    if s == "": return None, None
    m = PC_PAT.match(s)
    pc = m.group(1) if m else None
    store = None
    if pc is not None and pc in PC_TO_STORE:
        store = PC_TO_STORE[pc]
    else:
        sl = s.lower()
        for token, short in FALLBACK_STORE_TOKENS.items():
            if token in sl:
                store = short; break
    return pc, store

def ensure_pc_store(df: pd.DataFrame, df_raw_for_scan=None, log_prefix=""):
    df = df.copy()
    pc_col = find_col(df, [r"^pc\b", r"\bpc[\s_]*number\b", r"\bstore\s*#\b",
                           r"\bstore\s*number\b", r"\brestaurant\s*number\b", r"\blocation\s*id\b"])
    store_col = find_col(df, [r"^store\b", r"\bstore\s*name\b", r"\brestaurant\b",
                              r"\blocation\b", r"\blocation\s*name\b", r"\brevenue\s*center\b",
                              r"\bsite\b", r"\bunit\b", r"\bunit\s*name\b"])
    if pc_col is not None and "PC_Number" not in df.columns: df = df.rename(columns={pc_col:"PC_Number"})
    if store_col is not None and "Store" not in df.columns: df = df.rename(columns={store_col:"Store"})
    if "PC_Number" not in df.columns or "Store" not in df.columns:
        loc_col = find_col(df, [r"\blocation\s*name\b", r"\brevenue\s*center\b", r"\blocation\b",
                                r"\brestaurant\b", r"\bsite\b", r"\bunit\b", r"\bunit\s*name\b"])
        if loc_col is not None:
            pcs, stores = [], []
            for v in df[loc_col].tolist():
                pc, st = derive_pc_store_from_text(v)
                pcs.append(pc); stores.append(st)
            if "PC_Number" not in df.columns: df["PC_Number"] = pcs
            if "Store" not in df.columns: df["Store"] = stores
    need_pc = ("PC_Number" not in df.columns) or df["PC_Number"].isna().all()
    need_store = ("Store" not in df.columns) or df["Store"].isna().all()
    candidate = None
    if need_pc or need_store:
        for col in df.columns:
            for v in df[col].head(50):
                pc, st = derive_pc_store_from_text(v)
                if pc is not None:
                    candidate = (pc, st); break
            if candidate: break
        if not candidate and df_raw_for_scan is not None:
            for _, row in df_raw_for_scan.head(20).iterrows():
                for v in row:
                    pc, st = derive_pc_store_from_text(v)
                    if pc is not None:
                        candidate = (pc, st); break
                if candidate: break
        if candidate:
            pc, st = candidate
            if need_pc: df["PC_Number"] = pc
            if need_store: df["Store"] = st
    if "PC_Number" in df.columns:
        df["PC_Number"] = df["PC_Number"].astype(str).str.replace(".0","",regex=False).str.strip()
    if "Store" in df.columns:
        df["Store"] = df["Store"].astype(str).str.strip()
    log(f"{log_prefix}PC/Store present? PC_Number={ 'PC_Number' in df.columns } | Store={ 'Store' in df.columns }")
    return df

def ensure_date(df: pd.DataFrame, date_fallback, log_prefix=""):
    date_col = find_col(df, [r"^date\b", r"\bbusiness\s*date\b", r"\breport\s*date\b", r"\bperiod\b", r"\bday$"])
    if date_col is not None:
        df["Date"] = pd.to_datetime(df[date_col], errors="coerce").dt.date
    if "Date" not in df.columns or df["Date"].isna().all():
        df["Date"] = date_fallback
    log(f"{log_prefix}Date → {date_col if date_col is not None else f'filename {date_fallback}'}")
    return df

# ── Field mappers
def map_sales_summary(df):
    fields = {
        "Gross_Sales":[r"\bgross\s*sales\b"],
        "Net_Sales":[r"\bnet\s*sales\b"],
        "DD_Adjusted_No_Markup":[r"adjusted.*reportable.*(w/?o).*markup", r"adjusted.*w/o.*markup"],
        "PA_Sales_Tax":[r"\bsales\s*tax\b"],
        "DD_Discount":[r"\bdd\s*discount"],
        "Guest_Count":[r"\bguest\s*count\b"],
        "Avg_Check":[r"\bavg\s*check\b", r"\baverage\s*check\b"],
        "Gift_Card_Sales":[r"\bgift\s*card\s*sales\b"],
        "Void_Amount":[r"\bvoid\s*amount\b"],
        "Refund":[r"\brefunds?\b"],
        "Void_Qty":[r"\bvoid\s*qty\b", r"\bvoid\s*quantity\b"],
        "Paid_IN":[r"\bpaid\s*in\b"],
        "Paid_OUT":[r"\bpaid\s*out\b"],
        "Cash_IN":[r"\bcash\s*in\b"],
    }
    keep = ["Store","PC_Number","Date"]
    for out,pats in fields.items():
        col = find_col(df, pats)
        if col is not None: keep.append(col)
    sub = df[keep].copy()
    rename = {}
    for out,pats in fields.items():
        col = find_col(df, pats)
        if col is not None: rename[col] = out
    sub = sub.rename(columns=rename)
    for k in fields.keys():
        if k in sub.columns: sub[k] = sub[k].map(clean_num)
    return sub

def map_daypart(df):
    fields = {
        "Daypart":[r"^daypart$"],
        "Net_Sales":[r"\bnet\s*sales\b|\bsales\b"],
        "Percent_Sales":[r"(%|percent).*(sales)|% of sales"],
        "Check_Count":[r"\bcheck\s*count\b|\bguest\s*count\b|\bchecks?\b"],
        "Avg_Check":[r"\bavg\s*check\b|\baverage\s*check\b"],
    }
    keep = ["Store","PC_Number","Date"]
    for out,pats in fields.items():
        col = find_col(df, pats)
        if col is not None: keep.append(col)
    sub = df[keep].copy()
    rename = {}
    for out,pats in fields.items():
        col = find_col(df, pats)
        if col is not None: rename[col] = out
    sub = sub.rename(columns=rename)
    for k in ["Net_Sales","Percent_Sales","Check_Count","Avg_Check"]:
        if k in sub.columns: sub[k] = sub[k].map(clean_num)
    return sub

def map_subcategory(df):
    fields = {
        "Subcategory":[r"\bsubcategory\b|\bsubcategory name\b"],
        "Qty_Sold":[r"\bqty\b|\bquantity\b|\bqty sold\b"],
        "Net_Sales":[r"\bnet\s*sales\b|\bsales\b"],
        "Percent_Sales":[r"(%|percent).*(sales)|% sales|% of sales"],
    }
    keep = ["Store","PC_Number","Date"]
    for out,pats in fields.items():
        col = find_col(df, pats)
        if col is not None: keep.append(col)
    sub = df[keep].copy()
    rename = {}
    for out,pats in fields.items():
        col = find_col(df, pats)
        if col is not None: rename[col] = out
    sub = sub.rename(columns=rename)
    if "Subcategory" in sub.columns:
        sub = sub[~sub["Subcategory"].astype(str).str.lower().str.startswith("total")]
    for k in ["Qty_Sold","Net_Sales","Percent_Sales"]:
        if k in sub.columns: sub[k] = sub[k].map(clean_num)
    return sub

def map_tender(df):
    fields = {
        "Tender_Type":[r"\btender\b|\bcode\b|\bdesc(ription)?\b"],
        "Detail_Amount":[r"\bamount\b|\bdetail\s*amount\b|\bnet\b|\bsum of amount\b"],
    }
    keep = ["Store","PC_Number","Date"]
    for out,pats in fields.items():
        col = find_col(df, pats)
        if col is not None: keep.append(col)
    sub = df[keep].copy()
    rename = {}
    for out,pats in fields.items():
        col = find_col(df, pats)
        if col is not None: rename[col] = out
    sub = sub.rename(columns=rename)
    if "Detail_Amount" in sub.columns: sub["Detail_Amount"] = sub["Detail_Amount"].map(clean_num)
    tender_map = {"4000059":"Discover","4000061":"Visa","4000060":"Mastercard","4000058":"Amex",
                  "4000065":"GC Redeem","4000098":"Grub Hub","4000106":"Uber Eats","4000107":"Doordash"}
    if "Tender_Type" in sub.columns:
        sub["Tender_Type"] = sub["Tender_Type"].astype(str).str.strip().map(lambda x: tender_map.get(x.split()[0], x))
    return sub

def map_labor(df):
    fields = {
        "Labor_Position":[r"labor.*position|^position$|^labor position name$"],
        "Reg_Hours":[r"\breg(ular)?\s*hours?\b|straight\s*hours?"],
        "OT_Hours":[r"\bot\s*hours?\b|overtime\s*hours?"],
        "Total_Hours":[r"\btotal\s*hours?\b"],
        "Reg_Pay":[r"\breg(ular)?\s*pay\b|straight\s*pay\b"],
        "OT_Pay":[r"\bot\s*pay\b|overtime\s*pay\b"],
        "Total_Pay":[r"\btotal\s*pay\b"],
        "Percent_Labor":[r"% labor|percent.*labor|% of labor"],
    }
    keep = ["Store","PC_Number","Date"]
    for out,pats in fields.items():
        col = find_col(df, pats)
        if col is not None: keep.append(col)
    sub = df[keep].copy()
    rename = {}
    for out,pats in fields.items():
        col = find_col(df, pats)
        if col is not None: rename[col] = out
    sub = sub.rename(columns=rename)
    for k in ["Reg_Hours","OT_Hours","Total_Hours","Reg_Pay","OT_Pay","Total_Pay","Percent_Labor"]:
        if k in sub.columns: sub[k] = sub[k].map(clean_num)
    return sub

def map_sales_mix_detail(df):
    """
    Flexible mapper for item-level detail. We search for the most common headings
    we’ve seen in POS exports.
    """
    fields = {
        # category/subcategory (optional, keep if present)
        "Category":     [r"^category$", r"\bmajor\s*category\b", r"\bmenu\s*category\b"],
        "Subcategory":  [r"\bsubcategory\b|\bminor\s*category\b|\bsub\s*category\b"],
        # item identifiers
        "Item":         [r"^item$", r"\bitem\s*name\b", r"\bmenu\s*item\b", r"\bproduct\s*name\b", r"^product$"],
        # metrics
        "Qty_Sold":     [r"\bqty\b|\bqty\s*sold\b|\bquantity\b|\bqty\s*sold\b"],
        "Net_Sales":    [r"\bnet\s*sales\b|\bsales\b"],
        "Percent_Sales":[r"(%|percent).*(sales)|% sales|% of sales"],
        "Avg_Check":    [r"\bavg\s*check\b|\baverage\s*check\b"],  # sometimes present
    }
    keep = ["Store","PC_Number","Date"]
    for out,pats in fields.items():
        col = find_col(df, pats)
        if col is not None: keep.append(col)
    sub = df[keep].copy()
    rename = {}
    for out,pats in fields.items():
        col = find_col(df, pats)
        if col is not None: rename[col] = out
    sub = sub.rename(columns=rename)

    # Clean totals & numbers
    if "Item" in sub.columns:
        sub = sub[~sub["Item"].astype(str).str.lower().str.startswith("total")]
    for k in ["Qty_Sold","Net_Sales","Percent_Sales","Avg_Check"]:
        if k in sub.columns: sub[k] = sub[k].map(clean_num)
    return sub

MAPPER_BY_CATEGORY = {
    "sales_summary":        map_sales_summary,
    "sales_by_daypart":     map_daypart,
    "sales_by_subcategory": map_subcategory,
    "tender_type_metrics":  map_tender,
    "labor_metrics":        map_labor,
    "sales_by_order_type":  lambda df: df,          # not used by current inputs
    "sales_mix_detail":     map_sales_mix_detail,   # <— NEW
}

# ── Template schema loader (*_copy.xlsx) or default
def load_template_schema_for(raw_filename: str, category: str):
    candidates = sorted(list(RAW_DIR.glob("*_copy.xlsx")) + list(OUT_DIR.glob("*_copy.xlsx")))
    def first_sheet_cols(pp: Path):
        try:
            df = pd.read_excel(pp, sheet_name=0, nrows=0, engine="openpyxl")
            return list(df.columns)
        except Exception:
            return []
    for p in candidates:
        nm = norm(p.name).lower()
        if category == "sales_summary" and (("menu mix metrics" in nm) or ("sales mix metrics" in nm)): return first_sheet_cols(p)
        if category == "sales_by_daypart" and ("sales by daypart" in nm): return first_sheet_cols(p)
        if category == "sales_by_subcategory" and ("sales by subcategory" in nm): return first_sheet_cols(p)
        if category == "tender_type_metrics" and ("tender type" in nm): return first_sheet_cols(p)
        if category == "labor_metrics" and (("labor hours" in nm) or ("labor metrics" in nm)): return first_sheet_cols(p)
        if category == "sales_by_order_type" and ("order type" in nm): return first_sheet_cols(p)
        if category == "sales_mix_detail" and ("sales mix detail" in nm): return first_sheet_cols(p)
    return DEFAULT_SCHEMAS.get(category, [])

# ── Per-file processing
def process_one_input(path: Path):
    name = path.name
    name_low = norm(name).lower()
    print("\n=== Processing:", name, "===")

    category = category_from_filename(name_low)
    print("Category detected:", category)
    if not category:
        log(f"SKIP (no category): {name}")
        return False

    sheets = wb_sheetnames(path)
    print("Sheets found:", sheets)
    if not sheets:
        log(f"ERROR (no sheets): {name}")
        return False

    sheets_ordered = prefer_sheets_for_category(sheets, category)
    extra_terms = []
    if category == "sales_by_daypart": extra_terms = ["daypart"]
    elif category == "tender_type_metrics": extra_terms = ["tender","location name"]
    elif category == "labor_metrics": extra_terms = ["labor","position","hours"]
    elif category == "sales_by_subcategory": extra_terms = ["subcategory","qty"]
    elif category == "sales_summary": extra_terms = ["gross sales","net sales","revenue center"]
    elif category == "sales_mix_detail": extra_terms = ["item","product","qty","net sales","category"]

    df = None; chosen_sheet=None; header_row=None; hdr_score=None; df0_raw=None
    for sheet in sheets_ordered:
        try:
            dfx, hdr_idx, sc, df0 = read_with_header_detection(path, sheet, extra_terms)
            if dfx is not None and not dfx.dropna(how="all").empty:
                df = dfx; df0_raw = df0; chosen_sheet=sheet; header_row=hdr_idx; hdr_score=sc
                break
        except Exception as e:
            log(f"{name} sheet '{sheet}' read error: {e}")
            continue

    print("Chosen sheet:", chosen_sheet, "| header_row:", header_row, "| header_score:", hdr_score)
    if df is None:
        log(f"ERROR (no readable data): {name}")
        return False

    print("Detected columns (first 15):", list(df.columns[:15]))

    date_fb = filename_date_fallback(name)
    df = ensure_pc_store(df, df_raw_for_scan=df0_raw, log_prefix=f"[{category}] ")
    df = ensure_date(df, date_fb, log_prefix=f"[{category}] ")

    mapper = MAPPER_BY_CATEGORY.get(category, lambda d: d)
    mapped = mapper(df)
    print("Mapped columns now:", list(mapped.columns))

    schema_cols = load_template_schema_for(name, category)
    print("Template columns (target order):", schema_cols)
    if not schema_cols:
        log(f"{name} WARN: no template schema found; using defaults for {category}")
        schema_cols = DEFAULT_SCHEMAS[category]

    for c in schema_cols:
        if c not in mapped.columns:
            mapped[c] = pd.NA
    mapped = mapped[schema_cols].copy()

    # light numeric coercion by name hints
    num_hints = {"gross","net","avg","check","qty","quantity","tax","discount","amount","percent","guest","pay","hours","sales"}
    for c in mapped.columns:
        lc = c.lower()
        if any(k in lc for k in num_hints):
            mapped[c] = mapped[c].map(clean_num)

    out_name = name[:-5] + "_copy.xlsx" if name.lower().endswith(".xlsx") else name + "_copy.xlsx"
    out_path = OUT_DIR / out_name
    with pd.ExcelWriter(out_path) as xlw:
        mapped.to_excel(excel_writer=xlw, sheet_name="Sheet1", index=False)

    log(f"OK: {name} → sheet='{chosen_sheet}' hdr_row={header_row} hdr_score={hdr_score} → {out_path.name} (rows={len(mapped)}, cols={len(mapped.columns)})")
    print(f"WROTE: {out_path.name}  rows={len(mapped)}  cols={len(mapped.columns)}")
    return True

def main():
    files = []
    for p in sorted(list(RAW_DIR.glob("*.xls")) + list(RAW_DIR.glob("*.xlsx"))):
        nm = norm(p.name).lower()
        if "consolidated dunkin sales summary v2" in nm:
            files.append(p)

    if not files:
        msg = f"No consolidated files found in {RAW_DIR}"
        log(msg); print(msg); return

    print(f"Found {len(files)} consolidated files in {RAW_DIR}")
    ok, fail = 0, 0
    for path in files:
        try:
            if process_one_input(path): ok += 1
            else: fail += 1
        except Exception as e:
            log(f"FATAL {path.name}: {e}"); fail += 1

    print(f"Done. Wrote {ok} compiled file(s). Failed: {fail}.")
    print(f"Log: {LOG_FILE}")

if __name__ == "__main__":
    main()
